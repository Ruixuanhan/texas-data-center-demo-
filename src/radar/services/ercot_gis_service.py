"""ERCOT Generator Interconnection Status workbook ingestion for Project Radar.

The adapter reads the committed July 2026 GIS workbook, preserves each source row as
an immutable document, and creates normalized project records and evidence events.
It does not fabricate coordinates: ERCOT rows without a published project location
remain available in search/timelines but are not rendered as map markers.
"""

from __future__ import annotations

import hashlib
import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from radar.data.models import IngestionRun, Project, ProjectEvent, Signal, SourceDocument
from radar.intelligence.stage_rules import infer_ercot_stage


PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_ERCOT_WORKBOOK = PROJECT_ROOT / "data" / "fixtures" / "ercot_gis_july_2026.xlsx"
ERCOT_PRODUCT_URL = "https://www.ercot.com/mp/data-products/data-product-details?id=PG7-200-ER"
PARSER_VERSION = "ercot-gis-xlsx-1.0"
WORKBOOK_AS_OF = datetime(2026, 7, 31)

LARGE_SHEET = "Project Details - Large Gen"
SMALL_SHEET = "Project Details - Small Gen"


def _json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def _as_datetime(value: Any) -> datetime | None:
    return value if isinstance(value, datetime) else None


def _as_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _as_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        converted = float(value)
    except (TypeError, ValueError):
        return None
    # GIS rows may carry negative values for adjustments. Preserve those values in
    # the raw source payload but omit them from the current capacity projection.
    return converted if converted >= 0 else None


def _row_hash(payload: dict[str, Any]) -> str:
    return hashlib.sha256(json.dumps(payload, sort_keys=True, default=str).encode("utf-8")).hexdigest()


def _find_header_index(rows: list[tuple[Any, ...]]) -> int:
    for index, row in enumerate(rows):
        if "INR" in row and "Project Name" in row:
            return index
    raise ValueError("No ERCOT project header row was found")


def _power_type(fuel: str | None, technology: str | None, project_name: str | None) -> str:
    combined = " ".join(part for part in (fuel, technology, project_name) if part).lower()
    if "battery" in combined or "bess" in combined or technology == "BA":
        return "Battery"
    if "gas" in combined or technology in {"CC", "CT", "ST"}:
        return "Gas"
    if "solar" in combined or technology == "PV":
        return "Solar"
    if "wind" in combined or technology == "WT":
        return "Wind"
    if "nuclear" in combined:
        return "Nuclear"
    return _as_text(fuel) or "Unknown"


def _large_records(rows: list[tuple[Any, ...]]) -> Iterable[dict[str, Any]]:
    header_index = _find_header_index(rows)
    headers = list(rows[header_index])
    positions = {str(header): index for index, header in enumerate(headers) if header}
    for row in rows[header_index + 1:]:
        inr = row[positions["INR"]] if len(row) > positions["INR"] else None
        name = row[positions["Project Name"]] if len(row) > positions["Project Name"] else None
        if not inr or not name:
            continue
        payload = {header: _json_safe(row[index]) if index < len(row) else None for header, index in positions.items()}
        payload["size_category"] = "Large"
        yield payload


def _small_records(rows: list[tuple[Any, ...]]) -> Iterable[dict[str, Any]]:
    header_index = _find_header_index(rows)
    headers = list(rows[header_index])
    positions = {str(header): index for index, header in enumerate(headers) if header}
    for row in rows[header_index + 1:]:
        inr = row[positions["INR"]] if len(row) > positions["INR"] else None
        name = row[positions["Project Name"]] if len(row) > positions["Project Name"] else None
        if not inr or not name:
            continue
        payload = {header: _json_safe(row[index]) if index < len(row) else None for header, index in positions.items()}
        payload["size_category"] = "Small"
        yield payload


def load_ercot_gis_records(workbook_path: Path = DEFAULT_ERCOT_WORKBOOK) -> list[dict[str, Any]]:
    """Return normalized raw records from the GIS large- and small-generator tabs."""
    if not workbook_path.exists():
        raise FileNotFoundError(f"ERCOT GIS workbook not found: {workbook_path}")
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    records = list(_large_records(list(workbook[LARGE_SHEET].iter_rows(values_only=True))))
    records.extend(_small_records(list(workbook[SMALL_SHEET].iter_rows(values_only=True))))
    return records


def _source_stage(payload: dict[str, Any]) -> str:
    study_phase = _as_text(payload.get("GIM Study Phase"))
    if study_phase:
        return study_phase
    ia_signed = _as_text(payload.get("IA Signed"))
    if ia_signed:
        return f"IA Signed: {ia_signed}"
    return "GIS project record"


def ingest_ercot_gis(session: Session, workbook_path: Path = DEFAULT_ERCOT_WORKBOOK) -> IngestionRun:
    """Ingest a GIS workbook idempotently and emit source-backed project events."""
    workbook_bytes = workbook_path.read_bytes()
    run = IngestionRun(
        source="ERCOT GIS",
        status="running",
        artifact_hash=hashlib.sha256(workbook_bytes).hexdigest(),
        message=f"Importing {workbook_path.name}",
    )
    session.add(run)
    session.flush()

    records = load_ercot_gis_records(workbook_path)
    changed = 0
    for payload in records:
        inr = _as_text(payload.get("INR"))
        project_name = _as_text(payload.get("Project Name"))
        if not inr or not project_name:
            continue

        source_key = f"ercot-gis:{inr}"
        source_stage = _source_stage(payload)
        assessment = infer_ercot_stage(
            study_phase=_as_text(payload.get("GIM Study Phase")),
            ia_signed=_as_text(payload.get("IA Signed")),
            synchronization_date=_as_datetime(payload.get("Approved for Synchronization")),
        )
        project = session.scalar(select(Project).where(Project.source_project_key == source_key))
        previous_stage = project.radar_stage if project else None
        previous_confidence = project.stage_confidence if project else 0.0
        if project is None:
            project = Project(source_project_key=source_key, project_name=project_name)
            session.add(project)
            session.flush()

        project.project_name = project_name
        project.developer = _as_text(payload.get("Interconnecting Entity"))
        project.county = _as_text(payload.get("County"))
        project.estimated_mw = _as_float(payload.get("Capacity (MW)"))
        project.power_type = _power_type(
            _as_text(payload.get("Fuel")),
            _as_text(payload.get("Technology")),
            project_name,
        )
        project.source_stage = source_stage
        project.radar_stage = assessment.stage
        project.stage_confidence = assessment.confidence
        project.ercot_status = source_stage
        project.permit_status = "Not checked"
        project.latest_signal = f"ERCOT GIS July 2026: {assessment.rationale}"
        project.source = "ERCOT GIS"
        project.source_url = ERCOT_PRODUCT_URL
        project.source_updated_at = WORKBOOK_AS_OF
        session.flush()

        content_hash = _row_hash(payload)
        document = session.scalar(
            select(SourceDocument).where(
                SourceDocument.source == "ERCOT GIS",
                SourceDocument.content_hash == content_hash,
            )
        )
        if document is not None:
            continue

        document = SourceDocument(
            project_id=project.id,
            source="ERCOT GIS",
            source_url=ERCOT_PRODUCT_URL,
            published_at=WORKBOOK_AS_OF,
            content_hash=content_hash,
            parser_version=PARSER_VERSION,
            raw_payload=payload,
        )
        session.add(document)
        session.flush()
        signal = Signal(
            project_id=project.id,
            source_document_id=document.id,
            signal_type="ercot_gis_status",
            signal_text=f"ERCOT GIS INR {inr}: {source_stage}",
            occurred_at=WORKBOOK_AS_OF,
            confidence=assessment.confidence,
            metadata_json={"inr": inr, "radar_stage": assessment.stage, "rationale": assessment.rationale},
        )
        session.add(signal)

        event_type = "baseline_loaded" if previous_stage is None else "stage_changed" if previous_stage != assessment.stage else "evidence_added"
        title = (
            f"ERCOT GIS baseline: {project_name} is {assessment.stage}"
            if previous_stage is None
            else f"ERCOT GIS stage changed: {previous_stage} → {assessment.stage}"
            if previous_stage != assessment.stage
            else f"ERCOT GIS evidence updated: {project_name}"
        )
        session.add(
            ProjectEvent(
                project_id=project.id,
                event_type=event_type,
                title=title,
                detail=assessment.rationale,
                before_value=previous_stage,
                after_value=assessment.stage,
                confidence_delta=round(assessment.confidence - previous_confidence, 2),
                occurred_at=WORKBOOK_AS_OF,
                source_document_id=document.id,
            )
        )
        changed += 1

    run.status = "success"
    run.completed_at = datetime.now(UTC)
    run.records_seen = len(records)
    run.records_changed = changed
    run.message = f"Processed {len(records)} ERCOT GIS records; created {changed} new evidence events."
    return run
